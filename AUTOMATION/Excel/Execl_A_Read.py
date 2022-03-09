import openpyxl

loc="D:\\python\\nibsc.xlsx"

workbook=openpyxl.load_workbook(loc)

sheet=workbook["usermanagement"]

row=sheet.max_row

cell=sheet.max_column

df=sheet.cell(5,1).value="abc"

workbook.save("D:\\python\\nibsc.xlsx")

print(df)